"""
Email Extractor & Excel Export Tool for Substack Profiles

This script:
1. Reads existing substack_profiles.csv
2. Scans all columns for email patterns
3. Optionally scrapes profile pages for emails (mailto links, bio text)
4. Exports all data to a formatted Excel file (.xlsx)

Usage:
    python email_extractor.py                    # CSV scan only (fast)
    python email_extractor.py --scrape           # Also scrape profile pages (slower)
    python email_extractor.py --scrape --limit 50  # Scrape only 50 profiles
"""

import csv
import re
import argparse
import time
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Email regex pattern - matches most common email formats
EMAIL_PATTERN = re.compile(
    r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
    re.IGNORECASE
)


def extract_emails_from_text(text: str) -> list:
    """Extract all email addresses from a text string."""
    if not text:
        return []
    
    # Find all email matches
    emails = EMAIL_PATTERN.findall(text)
    
    # Clean and deduplicate
    cleaned = set()
    for email in emails:
        email = email.lower().strip()
        # Filter out obvious non-emails (image files, etc.)
        if not any(ext in email for ext in ['.png', '.jpg', '.gif', '.jpeg', '.svg', '.webp']):
            cleaned.add(email)
    
    return list(cleaned)


def scan_csv_for_emails(csv_path: str) -> dict:
    """
    Scan existing CSV file for email patterns in all columns.
    Returns dict mapping username to list of found emails.
    """
    emails_found = {}
    
    if not Path(csv_path).exists():
        print(f"❌ CSV file not found: {csv_path}")
        return emails_found
    
    print(f"📂 Scanning CSV for emails: {csv_path}")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        row_count = 0
        
        for row in reader:
            username = row.get('Username', '')
            found_emails = []
            
            # Scan all columns for email patterns
            for column, value in row.items():
                if value:
                    emails = extract_emails_from_text(value)
                    found_emails.extend(emails)
            
            if found_emails:
                emails_found[username] = list(set(found_emails))
            
            row_count += 1
    
    print(f"   ✅ Scanned {row_count} profiles")
    print(f"   📧 Found emails in {len(emails_found)} profiles from CSV data")
    
    return emails_found


def scrape_emails_from_profiles(csv_path: str, limit: int = 0) -> dict:
    """
    Scrape profile pages for email addresses.
    Looks for mailto: links and email patterns in page content.
    Returns dict mapping username to list of found emails.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("⚠️  Playwright not installed. Skipping profile scraping.")
        return {}
    
    emails_found = {}
    
    # Read profile URLs from CSV
    profiles = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            profiles.append({
                'username': row.get('Username', ''),
                'profile_url': row.get('Profile URL', '')
            })
    
    if limit > 0:
        profiles = profiles[:limit]
    
    print(f"\n🌐 Scraping {len(profiles)} profile pages for emails...")
    print("   (This may take a while due to rate limiting)")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        )
        page = context.new_page()
        
        for idx, profile in enumerate(profiles, 1):
            username = profile['username']
            url = profile['profile_url']
            
            if not url:
                continue
            
            try:
                # Progress indicator
                if idx % 10 == 0 or idx == 1:
                    print(f"   📄 Processing {idx}/{len(profiles)}: @{username}")
                
                page.goto(url, wait_until="domcontentloaded", timeout=15000)
                time.sleep(1)  # Wait for content to load
                
                # Extract page content
                page_content = page.content()
                
                # Look for mailto: links
                mailto_emails = page.evaluate("""
                    () => {
                        const mailtoLinks = document.querySelectorAll('a[href^="mailto:"]');
                        return Array.from(mailtoLinks).map(a => {
                            const href = a.getAttribute('href');
                            return href.replace('mailto:', '').split('?')[0];
                        });
                    }
                """)
                
                # Also extract emails from visible text content
                visible_text = page.evaluate("""
                    () => {
                        return document.body.innerText;
                    }
                """)
                
                text_emails = extract_emails_from_text(visible_text)
                
                # Combine all found emails
                all_emails = set(mailto_emails + text_emails)
                
                # Filter out substack.com emails (not useful for outreach)
                all_emails = [e for e in all_emails if 'substack.com' not in e.lower()]
                
                if all_emails:
                    emails_found[username] = list(all_emails)
                    print(f"      ✅ @{username}: Found {len(all_emails)} email(s)")
                
                # Rate limiting
                if idx < len(profiles):
                    time.sleep(0.5)
                    
            except Exception as e:
                # Silently continue on errors
                pass
        
        browser.close()
    
    print(f"\n   📧 Found emails in {len(emails_found)} profiles from scraping")
    
    return emails_found


def load_csv_data(csv_path: str) -> list:
    """Load all data from CSV file."""
    data = []
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(dict(row))
    
    return data


def export_to_excel(data: list, emails: dict, output_path: str):
    """
    Export profile data with emails to a formatted Excel file.
    """
    print(f"\n📊 Exporting to Excel: {output_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Substack Profiles"
    
    # Define columns (add Email as first column after Username)
    columns = [
        "Username",
        "Email",  # New column
        "Profile URL",
        "Subscribers",
        "Twitter",
        "Instagram",
        "TikTok",
        "LinkedIn",
        "Facebook",
        "YouTube",
        "Linktree",
        "Threads",
        "Bluesky",
        "GitHub",
        "Medium",
        "Other",
        "Category",
        "Scraped At"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    normal_font = Font(size=10)
    email_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Write headers
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Write data rows
    emails_added = 0
    for row_idx, profile in enumerate(data, 2):
        username = profile.get('Username', '')
        
        # Get email for this profile
        profile_emails = emails.get(username, [])
        email_str = ', '.join(profile_emails) if profile_emails else ''
        
        if email_str:
            emails_added += 1
        
        row_data = [
            username,
            email_str,
            profile.get('Profile URL', ''),
            profile.get('Subscribers', ''),
            profile.get('Twitter', ''),
            profile.get('Instagram', ''),
            profile.get('TikTok', ''),
            profile.get('LinkedIn', ''),
            profile.get('Facebook', ''),
            profile.get('YouTube', ''),
            profile.get('Linktree', ''),
            profile.get('Threads', ''),
            profile.get('Bluesky', ''),
            profile.get('GitHub', ''),
            profile.get('Medium', ''),
            profile.get('Other', ''),
            profile.get('Category', ''),
            profile.get('Scraped At', '')
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Highlight email cells with data
            if col_idx == 2 and value:  # Email column
                cell.fill = email_fill
                cell.font = Font(size=10, bold=True)
    
    # Auto-adjust column widths
    column_widths = {
        'Username': 20,
        'Email': 35,
        'Profile URL': 40,
        'Subscribers': 12,
        'Twitter': 35,
        'Instagram': 35,
        'TikTok': 30,
        'LinkedIn': 40,
        'Facebook': 40,
        'YouTube': 45,
        'Linktree': 30,
        'Threads': 30,
        'Bluesky': 35,
        'GitHub': 30,
        'Medium': 30,
        'Other': 45,
        'Category': 20,
        'Scraped At': 20
    }
    
    for col_idx, header in enumerate(columns, 1):
        width = column_widths.get(header, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Save workbook
    wb.save(output_path)
    
    print(f"   ✅ Exported {len(data)} profiles")
    print(f"   📧 {emails_added} profiles have emails")
    print(f"   💾 Saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Extract emails from Substack profiles and export to Excel'
    )
    parser.add_argument(
        '--scrape', 
        action='store_true',
        help='Also scrape profile pages for emails (slower but finds more)'
    )
    parser.add_argument(
        '--limit',
        type=int,
        default=0,
        help='Limit number of profiles to scrape (0 = all)'
    )
    parser.add_argument(
        '--input',
        type=str,
        default='substack_profiles.csv',
        help='Input CSV file path'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='substack_profiles_with_emails.xlsx',
        help='Output Excel file path'
    )
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("📧 EMAIL EXTRACTOR & EXCEL EXPORT")
    print("=" * 60)
    print(f"   Input:  {args.input}")
    print(f"   Output: {args.output}")
    print(f"   Scrape profiles: {'Yes' if args.scrape else 'No (CSV only)'}")
    if args.scrape and args.limit > 0:
        print(f"   Scrape limit: {args.limit} profiles")
    print()
    
    # Step 1: Scan CSV for emails
    csv_emails = scan_csv_for_emails(args.input)
    
    # Step 2: Optionally scrape profile pages
    scraped_emails = {}
    if args.scrape:
        scraped_emails = scrape_emails_from_profiles(args.input, args.limit)
    
    # Merge email results (scraped takes priority if different)
    all_emails = {}
    all_usernames = set(csv_emails.keys()) | set(scraped_emails.keys())
    
    for username in all_usernames:
        combined = set()
        if username in csv_emails:
            combined.update(csv_emails[username])
        if username in scraped_emails:
            combined.update(scraped_emails[username])
        if combined:
            all_emails[username] = list(combined)
    
    # Step 3: Load full CSV data
    data = load_csv_data(args.input)
    
    # Step 4: Export to Excel
    export_to_excel(data, all_emails, args.output)
    
    # Final summary
    print("\n" + "=" * 60)
    print("✅ COMPLETE")
    print("=" * 60)
    print(f"   Total profiles: {len(data)}")
    print(f"   Profiles with emails: {len(all_emails)}")
    print(f"   Excel file: {args.output}")
    
    if all_emails:
        print("\n📋 Sample emails found:")
        for i, (username, emails) in enumerate(list(all_emails.items())[:5]):
            print(f"   @{username}: {', '.join(emails)}")
        if len(all_emails) > 5:
            print(f"   ... and {len(all_emails) - 5} more")


if __name__ == "__main__":
    main()
